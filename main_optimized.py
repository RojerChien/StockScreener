import pandas as pd
import subprocess
import os
import sys
import time
import webbrowser
import requests
import datetime
import numpy as np
from highcharts import Highstock # pip install python-highcharts
from bs4 import BeautifulSoup # pip install beautifulsoup4
from yahooquery import Ticker
from openpyxl import load_workbook

pd.options.mode.chained_assignment = None  # 將警告訊息關閉

# start = 0
true_number = 0
# run_number = 0
today = str(datetime.datetime.now().date())


def get_ptp_tickers(url1='https://help.zh-tw.firstrade.com/article/841-new-1446-f-regulations',
                    url2='https://www.itigerup.com/bulletin/ptp'):
    # 爬取第一個網站的 PTP 股票代號列表
    response1 = requests.get(url1)
    soup1 = BeautifulSoup(response1.content, "html.parser")

    table1 = soup1.find("table", {"width": "100%"})
    tbody1 = table1.find("tbody")
    rows1 = tbody1.find_all("tr")
    ptp_lists = []

    for row in rows1:
        cols = row.find_all("td")
        if len(cols) == 3:
            symbol = cols[2].get_text().strip()
            ptp_lists.append(symbol)

    # 爬取第二個網站的 PTP 股票代號列表
    response2 = requests.get(url2)
    soup2 = BeautifulSoup(response2.text, 'html.parser')
    table2 = soup2.find('table', {'class': 'table'})
    rows2 = table2.find_all('tr')
    ptp_tickers = []

    for row in rows2:
        code = row.find('td')
        if code:
            ptp_tickers.append(code.text.strip())

    # 合併兩個列表，並取得不重覆的值，最終返回 PTP 股票代號列表
    ptp_lists.extend(ptp_tickers)
    ptp_lists = list(set(ptp_lists))
    ptp_lists_len = len(ptp_lists)
    print("Total PTP Tickers:", ptp_lists_len)
    return ptp_lists
    print(ptp_lists)


def remove_ptp_list(ptp_tickers, tickers):
    us_non_ptp_tickers = [x for x in tickers if x not in ptp_tickers]
    us_ptp_tickers = [x for x in tickers if x in ptp_tickers]
    print("PTP US stocks:", us_ptp_tickers)
    us_ptp_tickers_length = len(us_ptp_tickers)
    print("Total Removed PTP Tickers:", us_ptp_tickers_length)
    # print("Non-PTP US stocks:", us_non_ptp_tickers)
    return us_non_ptp_tickers


def add_vwap_to_chart(chart, dfin, window, color, line_width, yAxis=0, id=None):
    vwap = dfin[f'vwap{window}'].values.tolist()
    vwap = [
        [int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), round(v, 2)]
        for i, v in enumerate(vwap)
    ]
    chart.add_data_set(vwap, 'line', f'vwap{window}', yAxis=yAxis, color=color, lineWidth=line_width, id=id,
                       dataGrouping={'units': [['day', [1]]]})


def plot_highchart_chart(dfin, ticker_in, url, date=today):
    # 初始化highstock对象
    chart = Highstock(renderTo='container', width=None, height=930)  # 添加宽度和高度
    # chart = highstock(renderTo='container', width=1800, height=900)  # 添加宽度和高度

    # 加入RSI data
    rsi_data = dfin['RSI'].values.tolist()
    rsi_data = [
        [int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), round(r, 2)]
        for i, r in enumerate(rsi_data)
    ]

    # 將RSI加到數據中
    chart.add_data_set(rsi_data, 'line', 'RSI', yAxis=2, dataGrouping={'units': [['day', [1]]]})

    # 添加50日成交量移動平均線
    dfin['volume_50ma'] = dfin['volume'].rolling(window=50).mean()
    volume_50ma = dfin['volume_50ma'].values.tolist()
    volume_50ma = [
        [int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), v]
        for i, v in enumerate(volume_50ma)
    ]
    chart.add_data_set(volume_50ma, 'line', '50日成交量移動平均', yAxis=1, dataGrouping={'units': [['day', [1]]]})

    # 添加蜡烛图序列
    ohlc = dfin[['open', 'high', 'low', 'close']].values.tolist()
    ohlc = [[int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), round(o, 2), round(h, 2), round(l, 2), round(c, 2)] for
            i, (o, h, l, c) in enumerate(ohlc)]

    chart.add_data_set(ohlc, 'candlestick', ticker_in, dataGrouping={'units': [['day', [1]]]})
    add_vwap_to_chart(chart, dfin, 144, 'purple', 3, id='vwap144')
    add_vwap_to_chart(chart, dfin, 55, 'red', 3, id='vwap55')
    add_vwap_to_chart(chart, dfin, 21, 'orange', 3, id='vwap21')
    add_vwap_to_chart(chart, dfin, 5, 'blue', 3, id='vwap5')

    # 添加成交量序列
    volume = dfin[['open', 'close', 'volume']].values.tolist()
    volume = [{'x': int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), 'y': v, 'color': 'red' if o > c else 'green'} for
              i, (o, c, v) in enumerate(volume)]

    # chart.add_data_set(volume, 'column', '成交量', yAxis=1, dataGrouping={'units': [['day', [1]]]})
    # 禁用datagroup，讓volume在顯示時是正常的，但也可能影響效能
    chart.add_data_set(volume, 'column', '成交量', yAxis=1, dataGrouping={'enabled': False})

    # 设置图表选项
    options = {
        'rangeSelector': {
            'buttons': [{
                'type': 'month',
                'count': 1,
                'text': '1m'
            }, {
                'type': 'month',
                'count': 3,
                'text': '3m'
            }, {
                'type': 'month',
                'count': 6,
                'text': '6m'
            }, {
                'type': 'ytd',
                'text': 'YTD'
            }, {
                'type': 'year',
                'count': 1,
                'text': '1y'
            }, {
                'type': 'all',
                'text': 'All'
            }],
            'inputEnabled': True,
            'selected': 2  # 這裡的0代表上面buttons列表中的第一個按鈕，也就是'6m'
        },
        'chart': {
            'events': {
                'load': """
                    function () {
                        var chart = this;

                        // Create the toggleVwapLines function
                        var toggleVwapLines = function () {
                            var vwap144 = chart.get('vwap144'),
                                vwap55 = chart.get('vwap55'),
                                vwap21 = chart.get('vwap21');
                                vwap5 = chart.get('vwap5');
                            if (vwap144.visible) {
                                vwap144.hide();
                                vwap55.hide();
                                vwap21.hide();
                                vwap5.hide();
                            } else {
                                vwap144.show();
                                vwap55.show();
                                vwap21.show();
                                vwap5.show();
                            }
                        };

                        // Add a custom button
                        chart.renderer.button('Toggle VWAP Lines', null, null, toggleVwapLines)
                            .attr({
                                zIndex: 3
                            })
                            .add();
                    }
                """
            }
        },
        'navigation': {
            'buttonOptions': {
                'align': 'right',
                'verticalAlign': 'top',
                'y': 0
            }
        },
        'title': {'text': f'{ticker_in} ({date})'},
        'yAxis': [
            {'labels': {'align': 'right', 'x': -3},
             'title': {'text': 'OHLC'},
             'height': '60%',
             'lineWidth': 2},
            {'labels': {'align': 'right', 'x': -3},
             'title': {'text': 'Volume'},
             'top': '63%',
             'height': '20%',
             'offset': 0,
             'lineWidth': 2},
            {'labels': {'align': 'right', 'x': -3},
             'title': {'text': 'RSI'},
             'top': '86%',  # 調整此值以設置RSI圖表的位置
             'height': '14%',  # 調整此值以設置RSI圖表的高度
             'offset': 0,
             'lineWidth': 2,
             'plotLines': [
                 {'value': 30,
                  'color': '#FF4500',
                  'width': 1},
                 {'value': 70,
                  'color': '#FF4500',
                  'width': 1}
             ]}
        ],
        'tooltip': {
            'formatter': f"""
                function () {{
                    var dataIndex = this.points[0].point.index;
                    var s = '<b>' + Highcharts.dateFormat('%A, %b %e, %Y', this.x) + '</b>';
                    s += '<br/>';

                    this.points.forEach(function (point) {{
                        if (point.series.name === '{ticker_in}') {{
                            s += '<br/>' + point.series.name + ': ';
                            s += 'Open: ' + point.point.open.toFixed(2);
                            s += ', High: ' + point.point.high.toFixed(2);
                            s += ', Low: ' + point.point.low.toFixed(2);
                            s += ', Close: ' + point.point.close.toFixed(2);

                            // 計算漲跌幅
                            var change = 0;
                            if (dataIndex > 0) {{
                                var previousClose = point.series.options.data[dataIndex - 1][4];
                                change = ((point.point.close - previousClose) / previousClose) * 100;
                            }}
                            s += '<br/>漲跌幅: ' + change.toFixed(2) + '%';
                        }} else {{
                            s += '<br/>' + point.series.name + ': ' + point.y;
                        }}
                    }});

                    return s;
                }}
            """
        },
        'plotOptions': {
            'line': {
                'showInLegend': True
            },
            'candlestick': {
                'color': 'red',
                'upColor': 'green'
            },
            'column': {
                'borderColor': 'none'
            }
        },
        'subtitle': {
            'text': f'<a href="{url}" target="_blank" style="color: #003399; text-decoration: underline; cursor: pointer;">TradingView</a>',
            'useHTML': True,
            'align': 'center',
            'y': 35
        },
    }

    chart.set_dict_options(options)

    # 显示图表
    # chart.save_file('candlestick_volume')
    # with open('candlestick_volume.html', 'w', encoding='utf-8') as f:
    # filename = f'./HTML/{date}/{ticker_in}_{date}.html'
    filename = f'{ticker_in}_{date}.html'
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(chart.htmlcontent)
    filepath = os.path.abspath(filename)
    # 以下為開啟html的code，後續已經改成當程式執行完畢後，一次性的從dataframe去開啟所有html
    """if sys.platform == 'darwin':  # Mac OS X
        subprocess.run(['open', '-a', 'Google Chrome', filepath])
    elif sys.platform == 'win32':  # Windows
        subprocess.run(['start', 'chrome', filepath], shell=True)
    else:  # 其他平台，如 Linux
        subprocess.run(['xdg-open', filepath])"""
    return filepath
    # print(f'HTML File Name:{filepath}')
    # webbrowser.open(filepath)


def plot_highchart_chart_backup_20231108(dfin, ticker_in, url, date=today):
    # 初始化highstock对象
    chart = Highstock(renderTo='container', width=None, height=930)  # 添加宽度和高度
    # chart = highstock(renderTo='container', width=1800, height=900)  # 添加宽度和高度

    # 加入RSI data
    rsi_data = dfin['RSI'].values.tolist()
    rsi_data = [
        [int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), round(r, 2)]
        for i, r in enumerate(rsi_data)
    ]

    # 將RSI加到數據中
    chart.add_data_set(rsi_data, 'line', 'RSI', yAxis=2, dataGrouping={'units': [['day', [1]]]})

    # 添加50日成交量移動平均線
    dfin['volume_50ma'] = dfin['volume'].rolling(window=50).mean()
    volume_50ma = dfin['volume_50ma'].values.tolist()
    volume_50ma = [
        [int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), v]
        for i, v in enumerate(volume_50ma)
    ]
    chart.add_data_set(volume_50ma, 'line', '50日成交量移動平均', yAxis=1, dataGrouping={'units': [['day', [1]]]})

    # 添加蜡烛图序列
    ohlc = dfin[['open', 'high', 'low', 'close']].values.tolist()
    ohlc = [[int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), round(o, 2), round(h, 2), round(l, 2), round(c, 2)] for
            i, (o, h, l, c) in enumerate(ohlc)]

    chart.add_data_set(ohlc, 'candlestick', ticker_in, dataGrouping={'units': [['day', [1]]]})
    add_vwap_to_chart(chart, dfin, 144, 'purple', 3, id='vwap144')
    add_vwap_to_chart(chart, dfin, 55, 'red', 3, id='vwap55')
    add_vwap_to_chart(chart, dfin, 21, 'orange', 3, id='vwap21')
    add_vwap_to_chart(chart, dfin, 5, 'blue', 3, id='vwap5')

    # 添加成交量序列
    volume = dfin[['open', 'close', 'volume']].values.tolist()
    volume = [{'x': int(pd.Timestamp(dfin.index[i]).value // 10 ** 6), 'y': v, 'color': 'red' if o > c else 'green'} for
              i, (o, c, v) in enumerate(volume)]

    # chart.add_data_set(volume, 'column', '成交量', yAxis=1, dataGrouping={'units': [['day', [1]]]})
    # 禁用datagroup，讓volume在顯示時是正常的，但也可能影響效能
    chart.add_data_set(volume, 'column', '成交量', yAxis=1, dataGrouping={'enabled': False})

    # 设置图表选项
    options = {
        'chart': {
            'events': {
                'load': """
                    function () {
                        var chart = this;

                        // Create the toggleVwapLines function
                        var toggleVwapLines = function () {
                            var vwap144 = chart.get('vwap144'),
                                vwap55 = chart.get('vwap55'),
                                vwap21 = chart.get('vwap21');
                                vwap5 = chart.get('vwap5');
                            if (vwap144.visible) {
                                vwap144.hide();
                                vwap55.hide();
                                vwap21.hide();
                                vwap5.hide();
                            } else {
                                vwap144.show();
                                vwap55.show();
                                vwap21.show();
                                vwap5.show();
                            }
                        };

                        // Add a custom button
                        chart.renderer.button('Toggle VWAP Lines', null, null, toggleVwapLines)
                            .attr({
                                zIndex: 3
                            })
                            .add();
                    }
                """
            }
        },
        'navigation': {
            'buttonOptions': {
                'align': 'right',
                'verticalAlign': 'top',
                'y': 0
            }
        },
        'title': {'text': f'{ticker_in} ({date})'},
        'yAxis': [
            {'labels': {'align': 'right', 'x': -3},
             'title': {'text': 'OHLC'},
             'height': '70%',
             'lineWidth': 2},
            {'labels': {'align': 'right', 'x': -3},
             'title': {'text': '成交量'},
             'top': '72%',
             'height': '13%',
             'offset': 0,
             'lineWidth': 2},
            {'labels': {'align': 'right', 'x': -3},
             'title': {'text': 'RSI'},
             'top': '87%',  # 調整此值以設置RSI圖表的位置
             'height': '13%',  # 調整此值以設置RSI圖表的高度
             'offset': 0,
             'lineWidth': 2,
             'plotLines': [
                 {'value': 30,
                  'color': '#FF4500',
                  'width': 1},
                 {'value': 70,
                  'color': '#FF4500',
                  'width': 1}
             ]}
        ],
        'tooltip': {
            'formatter': f"""
                function () {{
                    var dataIndex = this.points[0].point.index;
                    var s = '<b>' + Highcharts.dateFormat('%A, %b %e, %Y', this.x) + '</b>';
                    s += '<br/>';

                    this.points.forEach(function (point) {{
                        if (point.series.name === '{ticker_in}') {{
                            s += '<br/>' + point.series.name + ': ';
                            s += 'Open: ' + point.point.open.toFixed(2);
                            s += ', High: ' + point.point.high.toFixed(2);
                            s += ', Low: ' + point.point.low.toFixed(2);
                            s += ', Close: ' + point.point.close.toFixed(2);

                            // 計算漲跌幅
                            var change = 0;
                            if (dataIndex > 0) {{
                                var previousClose = point.series.options.data[dataIndex - 1][4];
                                change = ((point.point.close - previousClose) / previousClose) * 100;
                            }}
                            s += '<br/>漲跌幅: ' + change.toFixed(2) + '%';
                        }} else {{
                            s += '<br/>' + point.series.name + ': ' + point.y;
                        }}
                    }});

                    return s;
                }}
            """
        },
        'plotOptions': {
            'line': {
                'showInLegend': True
            },
            'candlestick': {
                'color': 'red',
                'upColor': 'green'
            },
            'column': {
                'borderColor': 'none'
            }
        },
        'subtitle': {
            'text': f'<a href="{url}" target="_blank" style="color: #003399; text-decoration: underline; cursor: pointer;">TradingView</a>',
            'useHTML': True,
            'align': 'center',
            'y': 35
        },
    }

    chart.set_dict_options(options)

    # 显示图表
    # chart.save_file('candlestick_volume')
    # with open('candlestick_volume.html', 'w', encoding='utf-8') as f:
    # filename = f'./HTML/{date}/{ticker_in}_{date}.html'
    filename = f'{ticker_in}_{date}.html'
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(chart.htmlcontent)
    filepath = os.path.abspath(filename)
    # 以下為開啟html的code，後續已經改成當程式執行完畢後，一次性的從dataframe去開啟所有html
    """if sys.platform == 'darwin':  # Mac OS X
        subprocess.run(['open', '-a', 'Google Chrome', filepath])
    elif sys.platform == 'win32':  # Windows
        subprocess.run(['start', 'chrome', filepath], shell=True)
    else:  # 其他平台，如 Linux
        subprocess.run(['xdg-open', filepath])"""
    return filepath
    # print(f'HTML File Name:{filepath}')
    # webbrowser.open(filepath)


def open_html_files(df):
    for index, row in df.iterrows():
        # 檢查 highchart_html_path 是否不是 NA 且副檔名為 .html
        if pd.notna(row['highchart_html_path']) and row['highchart_html_path'].endswith('.html'):
            filepath = os.path.abspath(row['highchart_html_path'])
            if sys.platform == 'darwin':  # Mac OS X
                subprocess.run(['open', '-a', 'Google Chrome', filepath])
            elif sys.platform == 'win32':  # Windows
                subprocess.run(['start', 'chrome', filepath], shell=True)
            else:  # 其他平台，如 Linux
                subprocess.run(['xdg-open', filepath])

def calculate_rsi(data, period):
    delta = data['close'].diff().dropna()
    gain = delta.where(delta > 0, 0)
    loss = -delta.where(delta < 0, 0)

    avg_gain = gain.rolling(window=period).mean()
    avg_loss = loss.rolling(window=period).mean()

    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))

    return rsi


def get_us_tickers(category) -> dict:
    print(f'############ Start  Getting Ticker by new {category} mode')
    df = pd.read_excel("Tickers.xlsx", sheet_name=category)

    # 使用字典推导式创建字典
    us_tickers_dict = {
        str(sym).replace('/', '-'): {"Sector": sector, "Industry": industry}
        for sym, sector, industry in zip(df["Symbol"], df["Sector"], df["Industry"])
        if '^' not in str(sym)
    }

    tickers_length_pre = len(us_tickers_dict)
    print(f'Total {category} Tickers before remove PTP:', tickers_length_pre)

    ptp_lists = get_ptp_tickers()
    symbols_removed_ptp = [sym for sym in us_tickers_dict if sym not in ptp_lists]
    tickers_length_post = len(symbols_removed_ptp)
    print(f'Total {category} Tickers after remove PTP:', tickers_length_post)
    # print(us_tickers_dict)
    print(f'############ Finish Getting Ticker by new {category} mode')
    return us_tickers_dict, symbols_removed_ptp


def calculate_vwap(df_in, duration, ochl='close'):
    # data['Typical Price'] = (data['high'] + data['low'] + data['close']) / 3
    df_in['VWPrice'] = df_in[ochl] * df_in['volume']
    df_in[f'vwap{duration}'] = df_in['VWPrice'].rolling(window=duration).sum() / df_in['volume'].rolling(
        window=duration).sum()
    return df_in




def sel_yq_historical_data(data_all, ticker_in):
    # Select data from get_yq_financial_data by tickers
    # print(data_all)
    # print(ticker_in)
    # data_org = data_all.reset_index()

    # 確定資料裡面有ticker這個index的資料
    if ticker_in in data_all.index.get_level_values('symbol'):
        data_org = data_all.loc[ticker_in]
        data_org = calculate_vwap(data_org, 5)
        data_org = calculate_vwap(data_org, 8)
        data_org = calculate_vwap(data_org, 13)
        data_org = calculate_vwap(data_org, 21)
        data_org = calculate_vwap(data_org, 34)
        data_org = calculate_vwap(data_org, 55)
        data_org = calculate_vwap(data_org, 89)
        data_org = calculate_vwap(data_org, 144)
        data_org = calculate_vwap(data_org, 233)
        data_org = calculate_vwap(data_org, 377)
        data_org = calculate_vwap(data_org, 610)
        data_org['AvgVol'] = data_org['volume'].rolling(55).mean()  # 55為平均的天數
        data_org['RSI'] = calculate_rsi(data_org, rsi_period)
        # 确认数据点数量
        # print(len(data_org['volume']))
        # 检查是否有缺失值
        # print(data_org['volume'].isnull().sum())
        # 确认数据类型
        # print(data_org['volume'].dtype)
    else:
        print(f'Index not found:{ticker_in}')
        empty_df = pd.DataFrame()
        return empty_df
    # data_org = data_all.loc[ticker_in]
    # Check if the DataFrame has at least two rows
    # if data_org.shape[0] < 2:
    #     print("DataFrame must have at least two rows.")
    #     return 0
    # else:

    return data_org
    print("DataFrame has at least two rows.")


def get_tradingview_url(ticker_in, category):
    #  不是台灣的ticker
    if (category != "TW"):
        url = ("https://www.tradingview.com/chart/sWFIrRUP/?symbol=" + ticker_in)
    # 若是台灣的ticker，則會需要分上市及上櫃
    else:
        if (ticker_in[-1] == "W"):
            row2 = ticker_in[:4]
            url = ("https://www.tradingview.com/chart/sWFIrRUP/?symbol=TWSE%3A" + row2)
        else:
            row2 = ticker_in[:4]
            url = ("https://www.tradingview.com/chart/sWFIrRUP/?symbol=TPEX%3A" + row2)
    return url


def get_yq_financial_data(ticker_list):
    # print(ticker_list)
    get_data_start_time = time.time()  # Record start time
    print("Start Getting Stock Financial Data")
    # fin_data = Ticker(ticker_list).get_modules(['summaryDetail'])
    fin_data = Ticker(ticker_list, validate=True, progress=True).all_modules
    # fin_data = Ticker(ticker_list).get_modules(['incomeStatementHistoryQuarterly',
    # 'fundOwnership', 'summaryDetail', 'summaryProfile'])

    # print(fin_data)
    # fin_data = Ticker(ticker_list).get_modules(['summaryDetail', 'balanceSheetHistory'])
    get_data_end_time = time.time()  # Record end time
    get_data_elapsed_time = round(get_data_end_time - get_data_start_time, 2)  # Compute elapsed time
    print(f"Get Financial Data Elapsed time: {get_data_elapsed_time} seconds")
    # Save DataFrame as CSV
    # data_all.to_csv('data_all.csv', index=True)
    # Save DataFrame as Parquet
    # data_all.to_parquet('data_all.parquet', index=True)
    # print(fin_data)
    return fin_data


def sel_yq_financial_data(data_all, ticker_in):
    # 初始化变量
    industry = 'N/A'
    sector = 'N/A'

    # 提取键值为 'AA' 的数据
    data_aa = data_all.get(ticker_in, {})
    # print(f"data_aa: {data_aa}")
    # 检查提取的数据是否是字典
    if isinstance(data_aa, dict):
        # 提取 'industry' 的值
        industry = data_aa.get('assetProfile', {}).get('industry', 'N/A')

        # 提取 'sector' 的值
        sector = data_aa.get('assetProfile', {}).get('sector', 'N/A')

        print(f"Industry: {industry}")
        print(f"Sector: {sector}")
    else:
        print(f"Data for {ticker_in} is not a dictionary.")

    return industry, sector


def get_yq_historical_data(ticker_list):
    # print(ticker_list)
    get_data_start_time = time.time()  # Record start time
    print("############ Start  Getting Stock Historical Data")
    ticker = Ticker(ticker_list, asynchronous=True)
    data_all = ticker.history(period="3y", interval="1d")
    get_data_end_time = time.time()  # Record end time
    get_data_elapsed_time = round(get_data_end_time - get_data_start_time, 2)  # Compute elapsed time
    print(f"Get Data Elapsed time: {get_data_elapsed_time} seconds")
    print("############ Finish Getting Stock Historical Data")
    # Save DataFrame as CSV
    # data_all.to_csv('data_all.csv', index=True)
    # Save DataFrame as Parquet
    # data_all.to_parquet('data_all.parquet', index=True)
    return data_all


def check_vwap_scenario_old(scenario, last_vwap_value_144, vwap_condition):
    # vwap_condition is a boolean value for the specific scenario
    return "TRUE" if vwap_condition and last_vwap_value_144 else "FALSE"


def calculate_volatility(df, window=15):
    recent_closes = df['close'].tail(window)
    return recent_closes.max() / recent_closes.mean() - recent_closes.min() / recent_closes.mean()

def get_vwap_result(df, vwap_levels):
    result = True
    for i in range(len(vwap_levels) - 1):
        result &= df[f'vwap{vwap_levels[i]}'].iloc[-1] > df[f'vwap{vwap_levels[i + 1]}'].iloc[-1]
    return result

def check_vwap_scenario(scenario, last_vwap_value, vwap_increment):
    return "TRUE" if scenario in vwap_increment and last_vwap_value else "FALSE"


def get_vwap_result(df, vwap_levels):
    result = True
    for i in range(len(vwap_levels) - 1):
        result &= df[f'vwap{vwap_levels[i]}'].iloc[-1] > df[f'vwap{vwap_levels[i + 1]}'].iloc[-1]
    return result

def check_vwap_scenario(scenario, last_vwap_value, vwap_increment):
    return "TRUE" if scenario in vwap_increment and last_vwap_value else "FALSE"


def calculate_vwap_result(df, scenario):
    # Start with a series of True values, since we're going to use logical AND operations
    result = pd.Series(True, index=df.index)

    # Define the VWAP columns in the order they will be compared
    vwap_columns = ['vwap5', 'vwap21', 'vwap55', 'vwap89', 'vwap144', 'vwap233']

    # Find the index of the scenario in the vwap_columns list
    scenario_index = vwap_columns.index(f'vwap{scenario}')

    # Perform the comparison for each VWAP column up to the scenario
    for i in range(scenario_index):
        result &= (df[vwap_columns[i]] > df[vwap_columns[i + 1]])

    return result


def get_historical_data_for_ticker(ticker, update_historical_data, data_all_tickers):
    if update_historical_data:
        df = sel_yq_historical_data(data_all_tickers, ticker)
        df.to_csv('yq_historical.csv', index=False)
    else:
        if os.path.exists('yq_historical.csv'):
            df = pd.read_csv('yq_historical.csv')
    return df

def prepare_dataframe_with_calculations(df):
    # Add your calculations here, e.g., VWAP results, volume calculations, etc.
    return df


def should_analyze_stock(df, volatility_15, volatility):
    # Assuming there's a minimum threshold for 'avg_money_traded' that you want to check
    min_avg_money_traded = 2000000  # Example threshold
    return (
        len(df.index) > 144 and
        volatility_15 > 0.01 and
        volatility > 0.15 and
        df['avg_money_traded'].iloc[-1] >= min_avg_money_traded
    )


def update_output_dataframe(ticker, industry, sector, final_result, is_volume_greater_avg_volume_55,
                            volume_compare_wi_avg_volume_55, volatility_15, new_high_volume, volume_level,
                            price_near_high, should_analyze_stock, highchart_html_path, url,  df_in):

    new_row = {
    'Ticker': ticker,
    'Industry': industry,
    'Sector': sector,
    'isMatchVWAP': final_result,
    'isVolumeGreaterAvgVolume55': is_volume_greater_avg_volume_55,
    'volumeGrowthWithAvgVolume55': volume_compare_wi_avg_volume_55,
    'volatility15': volatility_15,
    'newHighVolume': new_high_volume,
    'volume_level': volume_level,
    'price_near_high': price_near_high,
    'should_analyze_stock': should_analyze_stock,
    'highchart_html_path': highchart_html_path,
    'url': url
    }
    # 將 new_row 轉換為 DataFrame
    new_row_df = pd.DataFrame([new_row])

    # 檢查 NA 值
    na_values = new_row_df.isna()

    # 如果存在 NA 值，打印出來
    if na_values.any().any():  # 第一個 any() 檢查每列是否有 NA，第二個 any() 檢查是否有任何列包含 NA
        na_columns = na_values.any()
        na_columns = na_columns[na_columns].index.tolist()  # 獲取包含 NA 值的列名列表
        print(f"NA values found in columns: {na_columns}")
        print(new_row_df[na_columns])  # 打印出包含 NA 值的列
    # 填充 NA 值，例如用 0 填充
    df_in = df_in.dropna(axis=1, how='all')
    new_row_df = new_row_df.dropna(axis=1, how='all')
    # df_out = pd.concat([df_in, pd.DataFrame([new_row])], ignore_index=True)
    df_out = pd.concat([df_in, new_row_df], ignore_index=True)
    # df_out = df_in.append(new_row, ignore_index=True)
    return df_out

def save_results_to_excel(df_out):
    df_out.to_excel('Result.xlsx', index=False)
    add_hyperlinks_to_excel('Result.xlsx')

def add_hyperlinks_to_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    url_col = 11  # Assuming URL is in the 11th column
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=url_col)
        cell.hyperlink = cell.value
        cell.style = 'Hyperlink'
    wb.save(filename)

def calculate_volatility(df, n_days):
    # 選擇最後n_days天的數據
    hist = df.tail(n_days)

    # 計算最高價格/平均價格和最低價格/平均價格
    avg_close_price = hist['close'].mean()
    max_close_price = hist['close'].max()
    min_close_price = hist['close'].min()
    volatility_h = max_close_price / avg_close_price - 1
    volatility_l = avg_close_price / min_close_price - 1
    # print(f"Volativity high: {n_days} days {volatility_h}")
    # print(f"Volativity low:  {n_days} days {volatility_l}")
    volatility = round(volatility_h + volatility_l, 2)
    # print(f"Volativity: {n_days} days {volatility}")

    return volatility


def calculate_volatility_duration(df, end_day, start_day='0'):
    # 選擇從start_day到end_day之間的數據
    hist = df.iloc[start_day:end_day + 1]

    # 計算最高價格/平均價格和最低價格/平均價格
    avg_close_price = hist['close'].mean()
    max_close_price = hist['close'].max()
    min_close_price = hist['close'].min()
    volatility_h = max_close_price / avg_close_price - 1
    volatility_l = avg_close_price / min_close_price - 1
    # print(f"Volativity high: {start_day} to {end_day} {volatility_h}")
    # print(f"Volativity low:  {start_day} to {end_day} {volatility_l}")
    volatility = round(volatility_h + volatility_l, 2)
    # print(f"Volativity: {start_day} to {end_day} {volatility}")

    return volatility


def vwap_strategy_screener_in_range(tickers_in, tickers_dict_in, scenario=144, days=233, percentage=2, ticker_type="US",
                                    run_number=0, update_historical_data=1):
    df_out = pd.DataFrame(columns=['Ticker', 'Industry', 'Sector', 'isMatchVWAP', 'isVolumeGreaterAvgVolume55',
                                   'volumeGrowthWithAvgVolume55', 'volatility15', 'newHighVolume', 'volume_level',
                                   'price_near_high', 'url'])

    data_all_tickers = get_yq_historical_data(tickers_in)

    for ticker in tickers_in:
        sector = tickers_dict_in[ticker]["Sector"]
        industry = tickers_dict_in[ticker]["Industry"]
        # print(ticker_type, ticker, run_number)
        run_number += 1
        df = get_historical_data_for_ticker(ticker, update_historical_data, data_all_tickers)
        if df.empty:
            continue
        data_length = len(df['volume'])
        if data_length >= scenario:
            # Add by Rojer Chien
            # Part1
            last_volume = df['volume'].iloc[-1]
            last_volume_wi_factor = last_volume * vol_factor
            last_avg_volume_55 = round(df['AvgVol'].iloc[-1], 2)
            # print(f"last_volume_wi_factor:{last_volume_wi_factor}")
            # print(f"last_avg_volume_55:{last_avg_volume_55}")
            is_volume_greater_avg_volume_55 = last_volume_wi_factor > last_avg_volume_55

            # Part2
            # 最後成交量大於55天平均成交量的比例，正數為較高；負數為較低
            if np.isnan(last_volume_wi_factor) or np.isnan(last_avg_volume_55):
                print("Warning: NaN values encountered")
                volume_compare_wi_avg_volume_55 = np.nan  # Assign NaN or some default value
            elif last_avg_volume_55 == 0:
                print("Warning: Division by zero encountered")
                volume_compare_wi_avg_volume_55 = np.nan  # Assign NaN or some default value
            else:
                volume_compare_wi_avg_volume_55 = round((last_volume_wi_factor / last_avg_volume_55) - 1, 2)
            # volume_compare_wi_avg_volume_55 = round((last_volume_wi_factor / last_avg_volume_55) - 1, 2)

            # Part3
            max_volume_144 = df['volume'].rolling(window=144).max().iloc[-1]
            volume_level = round((last_volume_wi_factor / max_volume_144) - 1, 2)
            new_high_volume = "False"  # Give new_high_volume a default value
            if volume_level > 0:
                new_high_volume = "True"
                print(f"New High Level{new_high_volume}")

            # Part4
            max_price_144 = df['high'].rolling(window=144).max().iloc[-1]
            last_close = df['close'].iloc[-1]
            last_close_pect_max_144 = (last_close / max_price_144) - 1
            if last_close_pect_max_144 < 0.05:
                price_near_high = "True"
            else:
                price_near_high = "False"

            # Part5
            url = get_tradingview_url(ticker, ticker_type)

            # df = prepare_dataframe_with_calculations(df)
            volatility_15 = calculate_volatility(df, 15)
            volatility = calculate_volatility(df, len(df.index))

            vwap_levels = [5, 21, 55, 89, 144, 233]
            vwap_increment = {level: get_vwap_result(df, vwap_levels[:i+1]) for i, level in enumerate(vwap_levels)}
            df['VWAP_Result'] = calculate_vwap_result(df, scenario)
            final_result = check_vwap_scenario(scenario, df['VWAP_Result'].iloc[-1], vwap_increment)
            # Assuming 'avg_money_traded' is calculated as the product of average volume and last VWAP value
            df['average_volume'] = df['volume'].rolling(55).mean()  # 55為平均的天數
            last_average_volume = round(df['average_volume'].iloc[-1], 2)
            last_VWAP = round(df['vwap55'].iloc[-1], 2)

            # print(f"Last Average Volume: {last_average_volume}")
            # print(f"Last VWAP Result: {last_VWAP}")
            # print(f"Final Result: {final_result}")
            df['avg_money_traded'] = df['average_volume'].iloc[-1] * df['vwap55'].iloc[-1]
            should_analyze = should_analyze_stock(df, volatility_15, volatility)
            # print (f"Should Analyze:{should_analyze }")
            if final_result == 'TRUE' and should_analyze :
                highchart_html_path = plot_highchart_chart(df, ticker, url, ticker_type)
            else:
                final_result = "SKIP"
                highchart_html_path = "NA"
            print(f"{ticker_type}-{run_number} {ticker} {final_result}")

            df_out = update_output_dataframe(ticker, industry, sector, final_result, is_volume_greater_avg_volume_55,
                                    volume_compare_wi_avg_volume_55, volatility_15, new_high_volume, volume_level,
                                    price_near_high, should_analyze , highchart_html_path, url, df_out)
                # df_out = update_output_dataframe(df_out, run_number, ticker, industry, sector, final_result, df)
            """  
            else:
                print("Volume or volatility not meet, skip analysis!")
            """
        else:
            print(f"{ticker_type}-{run_number} {ticker} Data too less!")
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    # Print the DataFrame
    open_html_files(df_out)
    # print(df_out)
    pd.reset_option('display.max_rows')
    pd.reset_option('display.max_columns')
    save_results_to_excel(df_out)

    df_out['url'] = df_out['url'].apply(lambda x: f'<a href="{x}">Link</a>')
    # 將DataFrame轉換為HTML表格，並移除Pandas的索引列，escape=False防止HTML標籤被轉義
    html_table = df_out.to_html(index=False, escape=False)

    # HTML和JavaScript的模板
    html_template = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
    <meta charset="UTF-8">
    <title>VWAP Strategy Screener Results</title>
    <link rel="stylesheet" type="text/css" href="https://cdn.datatables.net/1.10.21/css/jquery.dataTables.css">
    <script type="text/javascript" charset="utf8" src="https://code.jquery.com/jquery-3.5.1.js"></script>
    <script type="text/javascript" charset="utf8" src="https://cdn.datatables.net/1.10.21/js/jquery.dataTables.js"></script>
    </head>
    <body>
    {html_table}
    <script>
    $(document).ready( function () {{
        $('table').DataTable();
    }});
    </script>
    </body>
    </html>
    """

    # 將HTML保存到文件
    with open('vwap_strategy_screener_results.html', 'w', encoding='utf-8') as f:
        f.write(html_template)


# VWAP Scenario
scenario = 144  # 21, 55, 89, 144, 233共5種

# RSI的日期範圍
rsi_period = 14

# plotly的圖檔大小
plotly_resolution = "high"

if plotly_resolution == "high":
    jpg_resolution = 1080
else:
    jpg_resolution = 800

# volume Factor
vol_factor = 1  # volume factor for VWAP_SCREENER

# screener要判斷的平均天數
screener_day = 8

# 強制寫出plotly，主要用來測試
ForcePLOT = 0
OnlyPLOT = False  # True / False

# 若程式執行至一半中斷，可以設定重新執行的位置
test_start = 0
us_start = 0
tw_start = 0
yf_start = 0
etf_start = 0
fin_start = 0

# 設定要執行的種類
# VCP_TEST = 0
run_vwap_strategy_screener_in_range = 1  # VWAP均線的策略，均線呈多頭排序時買入，且限價在最高價正負2%的股票

# 要執行的ticker種類
TEST = 0
if TEST == 1:
    US = 0
else:
    US = 1

# 將讀取到的資料轉換為 list，並存入 tickers  變數中
if US == 1:
    us_tickers_dict, us_tickers = get_us_tickers("US")
if TEST == 1:
    print(f"Mode:TEST")
    us_tickers_dict, us_tickers = get_us_tickers("US")
    us_tickers = us_tickers[:100]
    us_tickers_dict = {ticker: us_tickers_dict[ticker] for ticker in us_tickers}

"""if TEST == 1:
    test_tickers = ['STX', 'IT', 'GIII']
    test_tickers_dict = {
        "STX": {"Sector": "Technology", "Industry": "Electronics"},
        "IT": {"Sector": "Technology", "Industry": "Software"},
        "GIII": {"Sector": "Finance", "Industry": "Banking"}
    }
"""

start_time = time.time()  # 記錄開始時間

#############################################################################

# debug mode => 寫出較多的資料
# debugmode = 1
# 決定是否要執行screener
lets_party = 1
# 20231105

if lets_party == 1:
    vwap_strategy_screener_in_range(us_tickers, us_tickers_dict)
    # if TEST == 1:
    #     category = "TEST"
    #     vwap_strategy_screener_in_range(test_tickers, test_tickers_dict)

    # if US == 1:
    #     category = "US"
    #    vwap_strategy_screener_in_range(us_tickers, us_tickers_dict)

end_time = time.time()  # 記錄結束時間
elapsed_time = round(end_time - start_time, 2)  # 計算運行時間
print(f"Elapsed time: {elapsed_time} seconds")
